from __future__ import annotations

import io
import zipfile
from pathlib import Path
from typing import Iterable

import pandas as pd
from openpyxl import load_workbook


class ExcelReader:
    """
    Reader untuk format laporan harian pelayanan pasien.

    Karakteristik file:
    - informasi filter/report ada di baris 1-25
    - header tabel ada di baris 26
    - data mulai di baris 27
    - beberapa file export memiliki styles.xml yang tidak valid
      (contoh: color rgb="thin"), jadi workbook perlu dipatch dulu.
    """

    DEFAULT_HEADER_ROW = 26  # 1-based Excel row number

    def __init__(self, file_path: str, sheet_name: str | None = None):
        self.file_path = Path(file_path)
        self.sheet_name = sheet_name

    def _validate_file(self) -> None:
        if not self.file_path.exists():
            raise FileNotFoundError(f"File tidak ditemukan: {self.file_path}")

    def _patched_workbook_bytes(self) -> io.BytesIO:
        """
        Beberapa file export tidak bisa dibaca openpyxl karena styles.xml tidak valid.
        Di sini kita patch minimal agar workbook tetap dapat dibaca.
        """
        self._validate_file()

        raw_output = io.BytesIO()
        with zipfile.ZipFile(self.file_path, "r") as zin, zipfile.ZipFile(raw_output, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if item.filename == "xl/styles.xml":
                    text = data.decode("utf-8", errors="ignore")
                    # contoh invalid xml dari file sample:
                    # <color rgb="thin"/> -> bukan aRGB valid
                    text = text.replace('rgb="thin"', 'indexed="64"')
                    data = text.encode("utf-8")
                zout.writestr(item, data)

        raw_output.seek(0)
        return raw_output

    def _load_sheet_values(self) -> list[list[object]]:
        workbook_stream = self._patched_workbook_bytes()
        wb = load_workbook(workbook_stream, data_only=True, read_only=True)
        ws = wb[self.sheet_name] if self.sheet_name else wb[wb.sheetnames[0]]
        return list(ws.iter_rows(values_only=True))

    def read_raw(self) -> pd.DataFrame:
        rows = self._load_sheet_values()
        if not rows:
            raise ValueError("File Excel kosong atau tidak ada worksheet yang bisa dibaca.")
        return pd.DataFrame(rows)

    def detect_header_row(self, candidate_rows: Iterable[int] = (26, 25, 27, 1)) -> int:
        """
        Mengembalikan nomor baris Excel (1-based) untuk header.
        Default utama baris 26 sesuai format sample.
        """
        raw_df = self.read_raw()

        required_headers = {"Nama Pasien", "No. eRM", "NIK", "Jenis Kelamin", "Tgl.Lahir"}
        for excel_row in candidate_rows:
            idx = excel_row - 1
            if idx < 0 or idx >= len(raw_df):
                continue

            row_values = {
                str(v).strip() for v in raw_df.iloc[idx].tolist()
                if v is not None and str(v).strip() != ""
            }
            if required_headers.issubset(row_values):
                return excel_row

        raise ValueError("Header tabel pasien tidak ditemukan. Pastikan format Excel sesuai sample.")

    def read(self, header_row: int | None = None, drop_empty_rows: bool = True) -> pd.DataFrame:
        raw_df = self.read_raw()
        header_row = header_row or self.detect_header_row()
        header_idx = header_row - 1

        headers = [
            "" if v is None else str(v).strip()
            for v in raw_df.iloc[header_idx].tolist()
        ]

        df = raw_df.iloc[header_idx + 1 :].copy()
        df.columns = headers

        # buang kolom kosong tanpa header
        df = df.loc[:, [str(c).strip() != "" for c in df.columns]]

        if drop_empty_rows:
            df = df.dropna(how="all")

        df = df.reset_index(drop=True)

        if df.empty:
            raise ValueError("Tidak ada data pasien setelah header dibaca.")

        return df

    def read_slice(self, start_row: int = 0, start_col: int = 0, header_row: int | None = None) -> pd.DataFrame:
        df = self.read(header_row=header_row)
        data = df.iloc[start_row:, start_col:].reset_index(drop=True)
        if data.empty:
            raise ValueError("Slice Excel kosong. Cek start_row/start_col.")
        return data
